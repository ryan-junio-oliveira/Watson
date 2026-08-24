"""Adapter de XLSX (e XLS via openpyxl): preserva planilhas como tabelas."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Optional

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from ingestion.adapters.base import SourceAdapter
from ingestion.models import LoadedDocument, Section, Table, sha256_text


class XlsxAdapter(SourceAdapter):
    source_type = "xlsx"
    supported_extensions = {".xlsx", ".xls"}
    MAX_ROWS_PER_SHEET = 5000

    def __init__(self, logger: Optional[logging.Logger] = None):
        self.logger = logger

    def extract(self, filepath: Path) -> LoadedDocument:
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for Excel files. "
                "Install it with: pip install openpyxl"
            )

        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        try:
            sections: list = []
            tables: list = []
            parts: list = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sections.append(Section(heading=f"Planilha: {sheet_name}"))
                parts.append(f"# Planilha: {sheet_name}")

                headers: list = []
                rows: list = []
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx >= self.MAX_ROWS_PER_SHEET:
                        parts.append("... (truncado)")
                        break
                    values = [str(v) if v is not None else "" for v in row]
                    if row_idx == 0:
                        headers = values
                    rows.append(values)
                    parts.append(" | ".join(values))

                if headers or rows:
                    tables.append(
                        Table(
                            section=sheet_name,
                            headers=headers,
                            rows=rows[1:] if headers else rows,
                            markdown=_to_markdown(headers, rows),
                        )
                    )

            content = "\n\n".join(parts)
            return LoadedDocument(
                content=content,
                filepath=str(filepath),
                filename=filepath.name,
                file_type=filepath.suffix.lower(),
                modified_at="",
                file_size=0,
                source_type=self.source_type,
                source_id=str(filepath),
                metadata={"sheets": len(wb.sheetnames), "tables": len(tables)},
                sections=sections,
                tables=tables,
                content_hash=sha256_text(content),
            )
        finally:
            wb.close()


def _to_markdown(headers, rows):
    lines = []
    if headers:
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "---|" * len(headers))
    for row in rows[1:]:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)
